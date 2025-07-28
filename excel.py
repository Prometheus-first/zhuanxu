import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取文件
file1 = pd.ExcelFile('7.14注册.xlsx')
file2 = pd.ExcelFile('714注册用户.xlsx')

# 获取所有表名
sheet_names1 = file1.sheet_names
sheet_names2 = file2.sheet_names

# 假设每个文件只有一个工作表，获取数据框
if sheet_names1:
    df1 = file1.parse(sheet_names1[0])
if sheet_names2:
    df2 = file2.parse(sheet_names2[0])

# 假设两个表中手机号码列的列名分别为'号码特征'和'手机号'
if '账号' in df1.columns and '手机号' in df2.columns:
    # 在6.23注册.xlsx中添加新列并标记匹配行
    df1['匹配标签'] = df1['账号'].isin(df2['手机号']).map({True: '保留', False: ''})

    # 在623注册用户.xlsx中添加新列并标记匹配行
    df2['处理标签'] = df2['手机号'].isin(df1['账号']).map({True: '假数据不处理', False: ''})

    # 保存6.23注册.xlsx更新后的数据到【已匹配】6.23注册.xlsx，保留格式
    wb1 = load_workbook('7.14注册.xlsx')
    ws1 = wb1.active
    for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    wb1.save('【已匹配】7.14注册.xlsx')

    # 保存623注册用户.xlsx更新后的数据
    df2.to_excel('714注册用户.xlsx', index=False)
else:
    print('未找到指定的列名，请检查列名是否正确。')